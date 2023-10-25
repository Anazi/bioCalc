from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import logging


class ExcelReport:

    def __init__(self, ttest_model, user_input_model):
        self.ttest_model = ttest_model
        self.user_input_model = user_input_model
        self.report_options = self.user_input_model.reporting_options
        self.sheet_names = self.report_options.sheet_names
        self.logger = logging.getLogger(__name__)

    def _generate_new_file_path(self):
        # Get the current timestamp
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

        # Break down the original file path into directory, filename, and extension
        dir_path, full_filename = os.path.split(self.report_options.file_path)
        filename, file_ext = os.path.splitext(full_filename)

        # Construct the new filename with a timestamp
        new_filename = f"{filename}_{timestamp}{file_ext}"

        return os.path.join(dir_path, new_filename)

    def save_to_excel(self, fig=None):
        try:
            new_file_path = self._generate_new_file_path()

            # Save the DataFrame to an Excel file
            self.ttest_model.solutions_df.to_excel(new_file_path,
                                                   sheet_name=self.sheet_names.data_sheet, index=False)

            if fig:
                temp_path = f"temp_plot.png"
                fig.savefig(temp_path)

                book = load_workbook(new_file_path)
                if self.sheet_names.result_sheet not in book.sheetnames:
                    book.create_sheet(self.sheet_names.result_sheet)
                sheet = book[self.sheet_names.result_sheet]

                img = Image(temp_path)
                sheet.add_image(img, 'E5')  # 'E5' is the cell where the image will be placed

                book.save(new_file_path)
                os.remove(temp_path)

                self.logger.info(f"Saved the graph to {new_file_path}.")

        except Exception as e:
            self.logger.error(f"An error occurred while saving to Excel: {e}")
